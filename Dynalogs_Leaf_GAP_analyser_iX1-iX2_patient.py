# -*- coding: utf-8 -*-

# Analyse of the GAP between expected and real leaf position in RapidArc (R1-R2) for bench A and B
# Author : Aurélien Corroyer-Dulmont
# Version : 22 march 2020

# Update 09/03/2020 : all the interesting results are automatically upload to the excel file Alex made. To do that pandas, openpyxl and xlsxwriter are used
# Update 09/03/2020 : ask for new measurement   
# Update 12/03/2020 : will look at the folder in "Z:/qualité" and performed automatically the analysis to all the dynalogs files present in the folder, it will also copy all the results in the excel file and activate VBA macro to archive the results
# Update 22/05/2020 : fit to the new excel files (one per RapidArc), change the theshold value for conformity (based on the previous 6 months results, using mean+1.96SD)
# Update 04/06/2020 : transfert the dynalog files into the correct folders (the one for patient and the other on for PFRTOAT, will delete others)

import datetime
import codecs
import statistics
import pandas as pad 
from openpyxl import load_workbook
import win32com.client
from path import Path
import shutil
import os

date = datetime.datetime.now()

def Dynalogs_Leaf_GAP_analyser(RapidName, filepath):

	### Get treatment date and PatientID with the name of the dynalog file ###
	LenghtFilePath = len(filepath)
	AcquisitionDate = filepath[LenghtFilePath-28:LenghtFilePath-20]
	AcquisitionDateExcel = str(AcquisitionDate[6:]) + "/" + str(AcquisitionDate[4:6]) + "/" + str(AcquisitionDate[:4])
	Banc = filepath[LenghtFilePath-29]
	IDPatient = filepath[LenghtFilePath-13:LenghtFilePath-4]
	MachineName = str(RapidName)

	savepath = "Z:/Aurelien_Dynalogs/Results_Analyses_Dynalogs/Patients_traités/Results_iX1_iX2/Dynalogs_analyser_results_ID" + str(IDPatient) + "_" + str(MachineName) + "_" + str(AcquisitionDate[6:]) + str(AcquisitionDate[4:6]) + str(AcquisitionDate[:4]) + ".txt"
	

	### Analyse the number of line which is different from dynalog file from an other ###
	file = open(filepath, 'r')

	LineNotEmpty = [1]
	LineCount = 0

	while (not LineNotEmpty) != True:
		LineNotEmpty = file.readlines(1)
		LineCount += 1

	LineCount -= 1 # because the loop goes at the end + 1 line
	LineCount -= 6 # six first line are information on ARC not dynalog measure
	file.close()

	
	file = open(filepath, 'r')

	i = 0
	for i in range(6):
		file.readlines(1)   
		
	# déclaration des list qui contiendront les valeurs des positions de lames attendues et réelles et celle contenant la différence (en mm) entre chaque lame
	ExpPosLeaf = []
	RealPosLeaf = []
	LeafGAP = []

	i = 0
	j = 0
	k = 0
	for i in range(LineCount): #boucle sur l'ensemble du nombre de ligne du fichier dynalog
		LineRaw = file.readlines(1) #première ligne d'intérêt avec les valeurs de positions de lames, mais copie dans une list tout le premier paragraphe, donc pas utilisable
		LineListStr = ",".join(LineRaw) # passage en mode string pour ensuite repasser en mode list mais en séparant les chiffres par la virgule de façon à avoir une liste avec une donnée par ...je ne connais pas le terme
		LineListGood = LineListStr.split(",") #permet d'avoir les valeurs du premier paragraphe dans une list
		
		for j in range(60):
			ExpPosLeaf.append(LineListGood[14+k])
			RealPosLeaf.append(LineListGood[15+k])
			k += 4
			j += 1
		k = 0

	# Boucle pour récupérer la différence entre les positions attendues et réelles (en 100ème de mm)
	m = 0
	IndexRepLame = LineCount*60 #car 60 lames
	for m in range(IndexRepLame):
		Difference = int(ExpPosLeaf[m]) - int(RealPosLeaf[m])
		LeafGAP.append(abs(Difference))

	# détermination de la différence max entre position attendues et réelles et ceci pour le banc de lame A et/ou B sachant que les positions seront positives ou négatives en fonction du banc de lames
	MaximumDifference = max(LeafGAP)

	MaxDifference = MaximumDifference/100

	#####création d'une table où les multiples des positions de lames sont présentes (dans le but de remonter au n° de lame défectueuse)
	TableLame = []
	MaxLeafGAP = []
	MeanLeafGAP = []
	SDLeafGAP = []
	LeafGAPTemp = []
	i = 0
	j = 0
	k = 0
	for k in range(60):
		for i in range(LineCount):
			TableLame.append(k+j)
			x = LeafGAP[k+j]
			LeafGAPTemp.append(x)
			j += 60
		MaxLeafGAP.append(max(LeafGAPTemp))
		MeanLeafGAP.append(statistics.mean(LeafGAPTemp))
		SDLeafGAP.append(statistics.stdev(LeafGAPTemp))
		LeafGAPTemp = []
		k += 1
		j = 0

	i=0
	for i in range(60):
		MeanLeafGAP[i] = round(MeanLeafGAP[i]/100, 3) # arrondit à 2 décimales et passage en mm
		SDLeafGAP[i] = round(SDLeafGAP[i]/100, 3)
		
	### calcul de la valeur moyenne maximale et de l'écart-type de ces moyennes ###
	MeanLeafGAPAllLeaf = max(MeanLeafGAP)
	SDLeafGAPAllLeaf = statistics.stdev(MeanLeafGAP)
	MeanLeafGAPAllLeaf = round(MeanLeafGAPAllLeaf, 3)
	SDLeafGAPAllLeaf = round(SDLeafGAPAllLeaf, 4)

	TableLameIndex = MaxLeafGAP.index(MaximumDifference) #on obtient donc l'indice dans la table qui est directement un multiple de 60 cad si <60 alors ce sera la lame n°1; si compris entre 1 et 2 alors lame n°2 etc etc...
	LameNumber = TableLameIndex+1 #car TableLameIndex renvoie un indice et non le n° de lame

	#########                               22th May 2020 update                                   ###########
	#########  Change conformity values (based on the first 6 months results, using mean+1.96SD)   ########### 
	### Analyse of the conformity, depending of the machine ###
	if MaxDifference < 1 or MeanLeafGAPAllLeaf < 0.1:
		ResultLeafGAP = "Conforme"
	else:
		ResultLeafGAP = "Hors tolérance"
		ResultMaxDifference = "HT"



	### Put the informations in the python's terminal ###
	if Banc == "A":
		print("POUR LE BANC A :\n")
		print(u"L'écart maximal entre la position attendue et la position réelle est de : " + str(MaxDifference) + " mm et ceci pour la lame n°" + str(LameNumber) +"\n")
		print(u"L'écart moyen maximal entre la position attendue et la position réelle est de : " + str(MeanLeafGAPAllLeaf) + " mm \n")
		print(u"L'écart-type entre ces moyennes est de : " + str(SDLeafGAPAllLeaf) + " mm \n")
		i = 0
		for i in range(len(MaxLeafGAP)): # Loop for max deviation of the 60 leafs
			print("Ecart maximal / moyen / SD pour la lame n°" + str(i+1) + ": " + str(MaxLeafGAP[i]/100) + " / " + str(MeanLeafGAP[i]) + " / " + str(SDLeafGAP[i]) + " mm")

			
	else:
		print("POUR LE BANC B :\n")
		print(u"L'écart maximal entre la position attendue et la position réelle est de : " + str(MaxDifference) + " mm et ceci pour la lame n°" + str(LameNumber) +"\n")
		print(u"L'écart moyen maximal entre la position attendue et la position réelle est de : " + str(MeanLeafGAPAllLeaf) + " mm \n")
		print(u"L'écart-type entre ces moyennes est de : " + str(SDLeafGAPAllLeaf) + " mm \n")
		i = 0
		for i in range(len(MaxLeafGAP)): # Loop for max deviation of the 60 leafs
			print("Ecart maximal / moyen / SD pour la lame n°" + str(i+1) + ": " + str(MaxLeafGAP[i]/100) + " / " + str(MeanLeafGAP[i]) + " / " + str(SDLeafGAP[i]) + " mm")


	print("\n")
	print(u"Le résultat du test est : " + str(ResultLeafGAP.upper()) +"\n\n")
	print(u"L'ensemble des résultats sont dans le dossier : \n" + str(savepath) +"\n\n")
	
	#### Craation and filling of the text file result ###
	filesave = open(savepath, 'a')
	filesave = codecs.open(savepath, 'a', encoding='Latin-1')     # Coding to include "é"
	filesave.write(u"Résultats de l'analyse des Dynalogs")
	filesave.write("\n\n")
	filesave.write("ID Patient : " + str(IDPatient))
	filesave.write("\n\n")
	filesave.write("Nom machine : " + str(MachineName))
	filesave.write("\n\n")
	filesave.write("Date d'acquisition : " + str(AcquisitionDate[6:]) + "/" + str(AcquisitionDate[4:6]) + "/" + str(AcquisitionDate[:4]))
	filesave.write("\n\n")
	filesave.write("Date d'analyse : " + str(date.day) + "/" + str(date.month) + "/" + str(date.year))
	filesave.write("\n\n")
	filesave.write("\n\n")
	if Banc == "A":
		filesave.write(u"Résultats de l'analyse des Dynalogs pour le banc A")
	else:
		filesave.write(u"Résultats de l'analyse des Dynalogs pour le banc B")
	filesave.write("\n\n")
	filesave.write("L'écart maximal entre la position attendue et la position réelle est de : " + str(MaxDifference) + " mm et ceci pour la lame n°" + str(LameNumber))
	filesave.write("\n\n")
	filesave.write(u"L'écart moyen entre la position attendue et la position réelle est de : " + str(MeanLeafGAPAllLeaf) + " mm")
	filesave.write("\n\n")
	filesave.write(u"L'écart-type entre ces moyennes est de : " + str(SDLeafGAPAllLeaf) + " mm")
	filesave.write("\n\n")
	filesave.write("Le résultat du test est : " + str(ResultLeafGAP.upper()))
	filesave.write("\n\n")

	i = 0
	for i in range(len(MaxLeafGAP)):
		filesave.write("Ecart maximal / moyen / SD pour la lame n°" + str(i+1) + ": " + str(MaxLeafGAP[i]/100) + " / " + str(MeanLeafGAP[i]) + " / " + str(SDLeafGAP[i]) + " mm")
		filesave.write("\n")  

	filesave.write("\n\n")
	filesave.close()
	
	### To know the number of leaf which is in trouble ###
	if ResultLeafGAP == "Hors tolérance":
		if Banc == "A":
			LameNumber = "A" + str(LameNumber)
		else:
			LameNumber = "B" + str(LameNumber)
	else:
		LameNumber = ""

	ListOfResults = []
	ListOfResults = [int(IDPatient), str(AcquisitionDateExcel), MaxDifference, str(LameNumber), MeanLeafGAPAllLeaf, SDLeafGAPAllLeaf, str(ResultLeafGAP)]
	
	return (ListOfResults)


def GetPatientInformation(IDPatient):
	""" To obtain information about plan's name, localisation of the tumor and gamma index and mean values; information obtained from DELTA4 excel file """
	
	df = pad.read_excel('Z:/1_CQ Patients/CQ Patients DELTA4 iX1-iX2.xlsm', sheet_name='DQA PATIENTS iX', usecols="A:D,O,P", nrows=1000, header=2)
	df = df.drop(0, axis=0)
	df = df.drop(1, axis=0)
	df = df.drop(2, axis=0)
	df = df.dropna()

	df_patient = df.loc[df["ID patient"]==IDPatient,:]
	try:
		plan_name = df_patient["Nom du plan de traitement"].array[0]
		localisation = df_patient["Localisation"].array[0]
	except:
		plan_name = ""
		localisation = ""
	try:
		gamma_index = df_patient["Gamma index"].array[0]
		gamma_moyen = df_patient["Gamma moyen"].array[0]
	except:
		gamma_index = "NaN"
		gamma_moyen = "NaN"

	PatientInformation = []
	PatientInformation = [int(IDPatient), plan_name, localisation, gamma_index, gamma_moyen]

	return PatientInformation


#########                               22th May 2020 update                                   ###########
#########                     Split to two excel files, one per RapidArc                       ###########

#########                               9th Mars 2020 update                                   ###########
######### Create Pandas Excel functions to upload the results in the excel data base file ###########
def ExportToExcel(MachineName, PatientInformation, ListOfResultsA, ListOfResultsB):
	ListOfResultsToExcel = ListOfResultsA[1:6]+ListOfResultsB[2:6]
	if MachineName == "RapidArc_iX_1":
		book = load_workbook('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/7_CLINAC iX 1/7-3 CQ -EN/7-1 CQ_quotidien/CQ quotidien Dynalog Patients_iX1.xlsm', read_only=False, keep_vba=True)
		writer = pad.ExcelWriter('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/7_CLINAC iX 1/7-3 CQ -EN/7-1 CQ_quotidien/CQ quotidien Dynalog Patients_iX1.xlsm', engine='openpyxl') 
		writer.book = book
		writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
		ws = writer.sheets['Dynalog_Patient_iX1']
		ws["D18"] = int(PatientInformation[0])
		ws["E18"] = str(PatientInformation[1])
		ws["F18"] = str(PatientInformation[2])
		ws["G18"] = float(PatientInformation[3])
		ws["H18"] = float(PatientInformation[4])
		ws["I18"] = str(ListOfResultsToExcel[0])
		ws["J18"] = float(ListOfResultsToExcel[1])
		ws["L18"] = str(ListOfResultsToExcel[2])
		ws["M18"] = float(ListOfResultsToExcel[3])
		ws["O18"] = float(ListOfResultsToExcel[4])
		ws["Q18"] = float(ListOfResultsToExcel[5])
		ws["R18"] = str(ListOfResultsToExcel[6])
		ws["T18"] = float(ListOfResultsToExcel[7])
		ws["V18"] = float(ListOfResultsToExcel[8])
		writer.save()
		xl = win32com.client.Dispatch('Excel.Application')
		xl.Workbooks.Open(Filename = '//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/7_CLINAC iX 1/7-3 CQ -EN/7-1 CQ_quotidien/CQ quotidien Dynalog Patients_iX1.xlsm', ReadOnly=1)  
		xl.Worksheets("Dynalog_Patient_iX1").Activate()
		xl.Application.Run("ArchiverDynalog")
		xl.Application.Quit()
		del xl

	if MachineName == "RapidArc_iX_2":
		book = load_workbook('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/10_CLINAC iX 2/10-3 CQ -EN/10-1_CQ_quotidien/CQ quotidien Dynalog Patients_iX2.xlsm', read_only=False, keep_vba=True)
		writer = pad.ExcelWriter('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/10_CLINAC iX 2/10-3 CQ -EN/10-1_CQ_quotidien/CQ quotidien Dynalog Patients_iX2.xlsm', engine='openpyxl') 
		writer.book = book
		writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
		ws = writer.sheets['Dynalog_Patient_iX2']
		ws["D18"] = int(PatientInformation[0])
		ws["E18"] = str(PatientInformation[1])
		ws["F18"] = str(PatientInformation[2])
		ws["G18"] = float(PatientInformation[3])
		ws["H18"] = float(PatientInformation[4])
		ws["I18"] = str(ListOfResultsToExcel[0])
		ws["J18"] = float(ListOfResultsToExcel[1])
		ws["L18"] = str(ListOfResultsToExcel[2])
		ws["M18"] = float(ListOfResultsToExcel[3])
		ws["O18"] = float(ListOfResultsToExcel[4])
		ws["Q18"] = float(ListOfResultsToExcel[5])
		ws["R18"] = str(ListOfResultsToExcel[6])
		ws["T18"] = float(ListOfResultsToExcel[7])
		ws["V18"] = float(ListOfResultsToExcel[8])
		writer.save()
		xl = win32com.client.Dispatch('Excel.Application')
		xl.Workbooks.Open(Filename = '//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/10_CLINAC iX 2/10-3 CQ -EN/10-1_CQ_quotidien/CQ quotidien Dynalog Patients_iX2.xlsm', ReadOnly=1)  
		xl.Worksheets("Dynalog_Patient_iX2").Activate()
		xl.Application.Run("ArchiverDynalog")
		xl.Application.Quit()
		del xl



#########                               12th Mars update                                   ###########
#########                            loop to analyse all the files in the folder           ###########
fileListR1 = []
for f in Path('Z:/Aurelien_Dynalogs/0000_Fichiers_Dynalogs_A_Analyser/Patients_traites/RapidArc_iX1').walkfiles(): 
	fileListR1.append(f)

newFileList = []
lastFileList = []
for i in range(len(fileListR1)):
	newFileList.append(fileListR1[i].replace('Path(',''))
	lastFileList.append(newFileList[i].replace('\\','/'))

dynalogFileListR1 = lastFileList

fileListR2 = []
for f in Path('Z:/Aurelien_Dynalogs/0000_Fichiers_Dynalogs_A_Analyser/Patients_traites/RapidArc_iX2').walkfiles(): 
	fileListR2.append(f)

newFileList = []
lastFileList = []
for i in range(len(fileListR2)):
	newFileList.append(fileListR2[i].replace('Path(',''))
	lastFileList.append(newFileList[i].replace('\\','/'))

dynalogFileListR2 = lastFileList

print("\n\nIl y a " + str(len(dynalogFileListR1)) + " fichiers dynalogs à analyser pour le RapidArc_iX1")
print("\n\nIl y a " + str(len(dynalogFileListR2)) + " fichiers dynalogs à analyser pour le RapidArc_iX2")

if len(dynalogFileListR1) != 0 or len(dynalogFileListR2) != 0:
	print("\n\nLancement du programme d'analyse\n\n")
	for i in range(int(len(dynalogFileListR1)/2)):
		ListOfResultsA = Dynalogs_Leaf_GAP_analyser("RapidArc_iX_1", str(dynalogFileListR1[i]))
		ListOfResultsB = Dynalogs_Leaf_GAP_analyser("RapidArc_iX_1", str(dynalogFileListR1[i+int(len(dynalogFileListR1)/2)]))
		PatientInformation = GetPatientInformation(ListOfResultsA[0])
		ExportToExcel("RapidArc_iX_1", PatientInformation, ListOfResultsA, ListOfResultsB)
	print("\n\nANALYSE DYNALOGS iX1 TERMINEE\n\n")

	for i in range(int(len(dynalogFileListR2)/2)):
		ListOfResultsA = Dynalogs_Leaf_GAP_analyser("RapidArc_iX_2", str(dynalogFileListR2[i]))
		ListOfResultsB = Dynalogs_Leaf_GAP_analyser("RapidArc_iX_2", str(dynalogFileListR2[i+int(len(dynalogFileListR2)/2)]))
		PatientInformation = GetPatientInformation(ListOfResultsA[0])
		ExportToExcel("RapidArc_iX_2", PatientInformation, ListOfResultsA, ListOfResultsB)  
	print("\n\nANALYSE DYNALOGS iX2 TERMINEE\n\n")
	
	#### Suppression des fichiers dynalogs analysés ###
	for file in dynalogFileListR1:
		shutil.rmtree(file)
	for file in dynalogFileListR2:
		shutil.rmtree(file)

	os.system("pause")