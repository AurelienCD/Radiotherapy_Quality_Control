# -*- coding: utf-8 -*-

# Analyse of the GAP between expected and real leaf position in RapidArc (R1-R2) for bench A and B
# Author : Aurélien Corroyer-Dulmont
# Version : 22 march 2020

# Update 09/03/2020 : all the interesting results are automatically upload to the excel file Alex made. To do that pandas, openpyxl and xlsxwriter are used
# Update 09/03/2020 : ask for new measurement   
# Update 12/03/2020 : will look at the folder in "Z:/qualité" and performed automatically the analysis to all the dynalogs files present in the folder, it will also copy all the results in the excel file and activate VBA macro to archive the results
# Update 22/05/2020 : fit to the new excel files (one per RapidArc), change the theshold value for conformity (based on the previous 6 months results, using mean+1.96SD)
# Update 04/06/2020 : transfert the dynalog files into the correct folders (the one for patient and the other on for PFRTOAT, will delete others)

import datetime
import codecs
import statistics
import pandas as pad 
from openpyxl import load_workbook
import win32com.client
from path import Path
import shutil
import os

date = datetime.datetime.now()

def Dynalogs_Leaf_GAP_analyser(filepath):

    ### Get treatment date and RapidArc number with the name of the dynalog file ###
    LenghtFilePath = len(filepath)
    AcquisitionDate = filepath[LenghtFilePath-38:LenghtFilePath-30]
    AcquisitionDateExcel = str(AcquisitionDate[6:]) + "/" + str(AcquisitionDate[4:6]) + "/" + str(AcquisitionDate[:4])
    Banc = filepath[LenghtFilePath-39]
    NumeroRapid =  filepath[LenghtFilePath-14:LenghtFilePath-13]
    MachineName = "RapidArc_iX_" + str(NumeroRapid)

    savepath = "Z:/Aurelien_Dynalogs/Results_Analyses_Dynalogs/LEAF_GAP_PFROTAT/Results_iX1_iX2/Dynalogs_LEAF_GAP_analyser_results_" + str(MachineName) + "_" + str(AcquisitionDate[6:]) + str(AcquisitionDate[4:6]) + str(AcquisitionDate[:4]) + ".txt"
    

    ### Analyse the number of line which is different from dynalog file from an other ###
    file = open(filepath, 'r')

    LineNotEmpty = [1]
    LineCount = 0

    while (not LineNotEmpty) != True:
       LineNotEmpty = file.readlines(1)
       LineCount += 1

    LineCount -= 1 # because the loop goes at the end + 1 line
    LineCount -= 6 # six first line are information on ARC not dynalog measure
    file.close()

    
    file = open(filepath, 'r')

    i = 0
    for i in range(6):
        file.readlines(1)   
        
    # déclaration des list qui contiendront les valeurs des positions de lames attendues et réelles et celle contenant la différence (en mm) entre chaque lame
    ExpPosLeaf = []
    RealPosLeaf = []
    LeafGAP = []

    i = 0
    j = 0
    k = 0
    for i in range(LineCount): #boucle sur l'ensemble du nombre de ligne du fichier dynalog
        LineRaw = file.readlines(1) #première ligne d'intérêt avec les valeurs de positions de lames, mais copie dans une list tout le premier paragraphe, donc pas utilisable
        LineListStr = ",".join(LineRaw) # passage en mode string pour ensuite repasser en mode list mais en séparant les chiffres par la virgule de façon à avoir une liste avec une donnée par ...je ne connais pas le terme
        LineListGood = LineListStr.split(",") #permet d'avoir les valeurs du premier paragraphe dans une list
        
        for j in range(60):
            ExpPosLeaf.append(LineListGood[14+k])
            RealPosLeaf.append(LineListGood[15+k])
            k += 4
            j += 1
        k = 0

    # Boucle pour récupérer la différence entre les positions attendues et réelles (en 100ème de mm)
    m = 0
    IndexRepLame = LineCount*60 #car 60 lames
    for m in range(IndexRepLame):
        Difference = int(ExpPosLeaf[m]) - int(RealPosLeaf[m])
        LeafGAP.append(abs(Difference))

    # détermination de la différence max entre position attendues et réelles et ceci pour le banc de lame A et/ou B sachant que les positions seront positives ou négatives en fonction du banc de lames
    MaximumDifference = max(LeafGAP)

    MaxDifference = MaximumDifference/100

    #####création d'une table où les multiples des positions de lames sont présentes (dans le but de remonter au n° de lame défectueuse)
    TableLame = []
    MaxLeafGAP = []
    MeanLeafGAP = []
    SDLeafGAP = []
    LeafGAPTemp = []
    i = 0
    j = 0
    k = 0
    for k in range(60):
        for i in range(LineCount):
            TableLame.append(k+j)
            x = LeafGAP[k+j]
            LeafGAPTemp.append(x)
            j += 60
        MaxLeafGAP.append(max(LeafGAPTemp))
        MeanLeafGAP.append(statistics.mean(LeafGAPTemp))
        SDLeafGAP.append(statistics.stdev(LeafGAPTemp))
        LeafGAPTemp = []
        k += 1
        j = 0

    i=0
    for i in range(60):
        MeanLeafGAP[i] = round(MeanLeafGAP[i]/100, 3) # arrondit à 2 décimales et passage en mm
        SDLeafGAP[i] = round(SDLeafGAP[i]/100, 3)
        
    ### calcul de la valeur moyenne maximale et de l'écart-type de ces moyennes ###
    MeanLeafGAPAllLeaf = max(MeanLeafGAP)
    SDLeafGAPAllLeaf = statistics.stdev(MeanLeafGAP)
    MeanLeafGAPAllLeaf = round(MeanLeafGAPAllLeaf, 3)
    SDLeafGAPAllLeaf = round(SDLeafGAPAllLeaf, 4)

    TableLameIndex = MaxLeafGAP.index(MaximumDifference) #on obtient donc l'indice dans la table qui est directement un multiple de 60 cad si <60 alors ce sera la lame n°1; si compris entre 1 et 2 alors lame n°2 etc etc...
    LameNumber = TableLameIndex+1 #car TableLameIndex renvoie un indice et non le n° de lame

    #########                               22th May 2020 update                                   ###########
    #########  Change conformity values (based on the first 6 months results, using mean+1.96SD)   ########### 
    ### Analyse of the conformity, depending of the machine ###
    if MachineName == "RapidArc_iX_1":
        if Banc == "A":
            if MaxDifference < 0.51 or MeanLeafGAPAllLeaf < 0.060:
                ResultLeafGAP = "Conforme"
            else:
                ResultLeafGAP = "Hors tolérance"
                ResultMaxDifference = "HT"
        else:
            if MaxDifference < 0.51 or MeanLeafGAPAllLeaf < 0.061:
                ResultLeafGAP = "Conforme"
            else:
                ResultLeafGAP = "Hors tolérance"
    
    if MachineName == "RapidArc_iX_2":
        if Banc == "A":
            if MaxDifference < 0.54 or MeanLeafGAPAllLeaf < 0.059:
                ResultLeafGAP = "Conforme"
            else:
                ResultLeafGAP = "Hors tolérance"
        else:
            if MaxDifference < 0.54 or MeanLeafGAPAllLeaf < 0.062:
                ResultLeafGAP = "Conforme"
            else:
                ResultLeafGAP = "Hors tolérance"


    ### Put the informations in the python's terminal ###
    if Banc == "A":
        print("POUR LE BANC A :\n")
        print(u"L'écart maximal entre la position attendue et la position réelle est de : " + str(MaxDifference) + " mm et ceci pour la lame n°" + str(LameNumber) +"\n")
        print(u"L'écart moyen maximal entre la position attendue et la position réelle est de : " + str(MeanLeafGAPAllLeaf) + " mm \n")
        print(u"L'écart-type entre ces moyennes est de : " + str(SDLeafGAPAllLeaf) + " mm \n")
        i = 0
        for i in range(len(MaxLeafGAP)): # Loop for max deviation of the 60 leafs
            print("Ecart maximal / moyen / SD pour la lame n°" + str(i+1) + ": " + str(MaxLeafGAP[i]/100) + " / " + str(MeanLeafGAP[i]) + " / " + str(SDLeafGAP[i]) + " mm")

            
    else:
        print("POUR LE BANC B :\n")
        print(u"L'écart maximal entre la position attendue et la position réelle est de : " + str(MaxDifference) + " mm et ceci pour la lame n°" + str(LameNumber) +"\n")
        print(u"L'écart moyen maximal entre la position attendue et la position réelle est de : " + str(MeanLeafGAPAllLeaf) + " mm \n")
        print(u"L'écart-type entre ces moyennes est de : " + str(SDLeafGAPAllLeaf) + " mm \n")
        i = 0
        for i in range(len(MaxLeafGAP)): # Loop for max deviation of the 60 leafs
            print("Ecart maximal / moyen / SD pour la lame n°" + str(i+1) + ": " + str(MaxLeafGAP[i]/100) + " / " + str(MeanLeafGAP[i]) + " / " + str(SDLeafGAP[i]) + " mm")


    print("\n")             
    print(u"Le résultat du test est : " + str(ResultLeafGAP.upper()) +"\n\n")
    print(u"L'ensemble des résultats sont dans le dossier : \n" + str(savepath) +"\n\n")
    
    #### Craation and filling of the text file result ###
    filesave = open(savepath, 'a')
    filesave = codecs.open(savepath, 'a', encoding='Latin-1')     # Coding to include "é"
    filesave.write(u"Résultats de l'analyse des Dynalogs")
    filesave.write("\n\n")
    filesave.write("Machine : " + str(MachineName))
    filesave.write("\n\n")
    filesave.write("Date d'acquisition : " + str(AcquisitionDate[6:]) + "/" + str(AcquisitionDate[4:6]) + "/" + str(AcquisitionDate[:4]))
    filesave.write("\n\n")
    filesave.write("Date d'analyse : " + str(date.day) + "/" + str(date.month) + "/" + str(date.year))
    filesave.write("\n\n")
    filesave.write("\n\n")
    if Banc == "A":
        filesave.write(u"Résultats de l'analyse des Dynalogs pour le banc A")
    else:
        filesave.write(u"Résultats de l'analyse des Dynalogs pour le banc B")
    filesave.write("\n\n")
    filesave.write("L'écart maximal entre la position attendue et la position réelle est de : " + str(MaxDifference) + " mm et ceci pour la lame n°" + str(LameNumber))
    filesave.write("\n\n")
    filesave.write(u"L'écart moyen entre la position attendue et la position réelle est de : " + str(MeanLeafGAPAllLeaf) + " mm")
    filesave.write("\n\n")
    filesave.write(u"L'écart-type entre ces moyennes est de : " + str(SDLeafGAPAllLeaf) + " mm")
    filesave.write("\n\n")
    filesave.write("Le résultat du test est : " + str(ResultLeafGAP.upper()))
    filesave.write("\n\n")

    i = 0
    for i in range(len(MaxLeafGAP)):
      filesave.write("Ecart maximal / moyen / SD pour la lame n°" + str(i+1) + ": " + str(MaxLeafGAP[i]/100) + " / " + str(MeanLeafGAP[i]) + " / " + str(SDLeafGAP[i]) + " mm")
      filesave.write("\n")  

    filesave.write("\n\n")
    filesave.close()
    
    ### For Dynalogs_Leaf_GAP_Watcher program, need 0 or 1 in text file to give the information to Solène and Cindy about the result of the analysis ###
    ### (strange approach due to the fact that they do not have python libraries) ###
    if ResultLeafGAP == "Hors tolérance":
        if MachineName == "RapidArc_iX_1":
            filesave = open("Z:/Aurelien_Dynalogs/Results_Analyses_Dynalogs/LEAF_GAP_PFROTAT/Results_iX1_iX2/temp_results_R1.txt", 'w')
            filesave.write("1")
            filesave.close()
        elif MachineName == "RapidArc_iX_2":
            filesave = open("Z:/Aurelien_Dynalogs/Results_Analyses_Dynalogs/LEAF_GAP_PFROTAT/Results_iX1_iX2/temp_results_R2.txt", 'w')
            filesave.write("1")
            filesave.close()

    else:
        filesaveR1 = open("Z:/Aurelien_Dynalogs/Results_Analyses_Dynalogs/LEAF_GAP_PFROTAT/Results_iX1_iX2/temp_results_R1.txt", 'w')
        filesaveR2 = open("Z:/Aurelien_Dynalogs/Results_Analyses_Dynalogs/LEAF_GAP_PFROTAT/Results_iX1_iX2/temp_results_R2.txt", 'w')
        filesaveR1.write("0")
        filesaveR2.write("0")
        filesaveR1.close()
        filesaveR2.close()

    ### To know the number of leaf which is in trouble ###
    if ResultLeafGAP == "Hors tolérance":
        if Banc == "A":
            LameNumber = "A" + str(LameNumber)
        else:
            LameNumber = "B" + str(LameNumber)
    else:
        LameNumber = ""

    ListOfResults = []
    ListOfResults = [MachineName, str(AcquisitionDateExcel), MaxDifference, str(LameNumber), MeanLeafGAPAllLeaf, SDLeafGAPAllLeaf, str(ResultLeafGAP)]
    
    return (ListOfResults)


#########                               22th May 2020 update                                   ###########
#########                     Split to two excel files, one per RapidArc                       ###########

#########                               9th Mars 2020 update                                   ###########
######### Create Pandas Excel functions to upload the results in the excel data base file ###########
def ExportToExcel(ListOfResultsA, ListOfResultsB):
    MachineName = str(ListOfResultsA[0])
    #Results_Analyses_Dynalogs = str(ListOfResultsA[5])
    #ListOfResultsA = ListOfResultsA[2:5]
    #ListOfResultsB = ListOfResultsB[2:5]
    ListOfResultsToExcel = ListOfResultsA[1:6]+ListOfResultsB[2:6]
    #ListOfResultsToExcel = [x.replace('.', ',') for x in ListOfResultsToExcel]
    #df = pad.DataFrame(ListOfResultsToExcel)
    if MachineName == "RapidArc_iX_1":
        book = load_workbook('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/7_CLINAC iX 1/7-3 CQ -EN/7-1 CQ_quotidien/CQ quotidien Dynalog PFROTAT - iX1.xlsm', read_only=False, keep_vba=True)
        writer = pad.ExcelWriter('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/7_CLINAC iX 1/7-3 CQ -EN/7-1 CQ_quotidien/CQ quotidien Dynalog PFROTAT - iX1.xlsm', engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        #df.to_excel(writer, "Dynalog_PFROTAT_iX1",startrow=9, startcol=0, header=False, index=False)
        ws = writer.sheets['Dynalog_PFROTAT_iX1']
        ws["F18"] = str(ListOfResultsToExcel[0])
        ws["G18"] = float(ListOfResultsToExcel[1])
        ws["I18"] = str(ListOfResultsToExcel[2])
        ws["J18"] = float(ListOfResultsToExcel[3])
        ws["L18"] = float(ListOfResultsToExcel[4])
        ws["N18"] = float(ListOfResultsToExcel[5])
        ws["P18"] = str(ListOfResultsToExcel[6])
        ws["Q18"] = float(ListOfResultsToExcel[7])
        ws["S18"] = float(ListOfResultsToExcel[8])
        writer.save()
        xl = win32com.client.Dispatch('Excel.Application')
        xl.Workbooks.Open(Filename = '//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/7_CLINAC iX 1/7-3 CQ -EN/7-1 CQ_quotidien/CQ quotidien Dynalog PFROTAT - iX1.xlsm', ReadOnly=1)  
        xl.Worksheets("Dynalog_PFROTAT_iX1").Activate()
        #ws = xl.ActiveSheet ############# ne sert à rien pour moi mais à tester si bug 
        xl.Application.Run("ArchiverDynalog_PROTAT_iX1")
        xl.Application.Quit()
        del xl

    else:
        book = load_workbook('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/10_CLINAC iX 2/10-3 CQ -EN/10-1_CQ_quotidien/CQ quotidien Dynalog PFROTAT - iX2.xlsm', read_only=False, keep_vba=True)
        writer = pad.ExcelWriter('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/10_CLINAC iX 2/10-3 CQ -EN/10-1_CQ_quotidien/CQ quotidien Dynalog PFROTAT - iX2.xlsm', engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        #df.to_excel(writer, "Dynalog_PFROTAT_iX2",startrow=9, startcol=0, header=False, index=False)
        ws = writer.sheets['Dynalog_PFROTAT_iX2']
        ws["F18"] = str(ListOfResultsToExcel[0])
        ws["G18"] = float(ListOfResultsToExcel[1])
        ws["I18"] = str(ListOfResultsToExcel[2])
        ws["J18"] = float(ListOfResultsToExcel[3])
        ws["L18"] = float(ListOfResultsToExcel[4])
        ws["N18"] = float(ListOfResultsToExcel[5])
        ws["P18"] = str(ListOfResultsToExcel[6])
        ws["Q18"] = float(ListOfResultsToExcel[7])
        ws["S18"] = float(ListOfResultsToExcel[8])
        writer.save()
        xl = win32com.client.Dispatch('Excel.Application')
        xl.Workbooks.Open(Filename = '//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/10_CLINAC iX 2/10-3 CQ -EN/10-1_CQ_quotidien/CQ quotidien Dynalog PFROTAT - iX2.xlsm', ReadOnly=1)  
        xl.Worksheets("Dynalog_PFROTAT_iX2").Activate()
        #ws = xl.ActiveSheet ############# ne sert à rien pour moi mais à tester si bug 
        xl.Application.Run("ArchiverDynalog_PROTAT_iX2")
        xl.Application.Quit()
        del xl
  


#########                               04th June update                                   ###########
#########            transfert the dynalog files into the correct folders                  ###########
#########   (the one for patient and the other on for PFRTOAT, will delete others)         ###########


PFRot_File_List = []
Patient_File_List = []
for elm in fileList:
    if elm[94:] == "PF ROTAT.dlg":
        PFRot_File_List.append(elm)
    else:
        Patient_File_List.append(elm)

TruePatient_File_List = []
Bin_File_List = []
for elm in Patient_File_List:
    try:
        float(elm[68:75]) > 10
        TruePatient_File_List.append(elm)
    except:
        Bin_File_List.append(elm)

for file in PFRot_File_List:
    shutil.move(file, 'Z:/Aurelien_Dynalogs/0000_Fichiers_Dynalogs_A_Analyser/PFRTOTAT_TOP')

for file in TruePatient_File_List:
    shutil.move(file, 'Z:/Aurelien_Dynalogs/0000_Fichiers_Dynalogs_A_Analyser/Patients traités')

for elm in Bin_File_List:
    os.remove(elm)



#########                               12th Mars update                                   ###########
#########                            loop to analyse all the files in the folder           ###########
fileList = []
for f in Path('Z:/Aurelien_Dynalogs/0000_Fichiers_Dynalogs_A_Analyser/PFRTOTAT_TOP').walkfiles(): 
    fileList.append(f)

newFileList = []
lastFileList = []
for i in range(len(fileList)):
    newFileList.append(fileList[i].replace('Path(',''))
    lastFileList.append(newFileList[i].replace('\\','/'))
    i += 1

dynalogFileList = lastFileList

print("\n\nIl y a " + str(len(dynalogFileList)) + " fichiers dynalogs à analyser")

if len(dynalogFileList) != 0:
    print("\n\nLancement du programme d'analyse\n\n")
    for i in range(int(len(dynalogFileList)/2)):
        ListOfResultsA = Dynalogs_Leaf_GAP_analyser(str(dynalogFileList[i]))
        ListOfResultsB = Dynalogs_Leaf_GAP_analyser(str(dynalogFileList[i+int(len(dynalogFileList)/2)]))
        ExportToExcel(ListOfResultsA, ListOfResultsB)
        i += 1
    
    print("\n\nANALYSE TERMINEE\n\n")
    #### Déplacement des fichiers dynalogs analysés dans un répertoire d'archive ###
    for file in dynalogFileList:
        shutil.move(file, 'Z:/Aurelien_Dynalogs/Results_Analyses_Dynalogs/LEAF_GAP_PFROTAT/Archives')

    ### Annonce des résultats de l'analyse ###
    Results_Dynalogs_AnalysisA = str(ListOfResultsA[6])
    Results_Dynalogs_AnalysisB = str(ListOfResultsB[6])

    if Results_Dynalogs_AnalysisA == "Hors tolérance" or Results_Dynalogs_AnalysisB == "Hors tolérance":
        print("\n\nRESULTATS NON CONFORMES\n\n")
        MachineName = str(ListOfResultsA[0])
        if MachineName == "RapidArc_iX_1":
            os.startfile('\\\\s-grp\\grp\\RADIOPHY\\Contrôle Qualité RTE\\Contrôle Qualité RTE-accélérateurs\\7_CLINAC iX 1\\7-3 CQ -EN\\7-1 CQ_quotidien\\CQ quotidien Dynalog PFROTAT - iX1.xlsm')
            os.system("pause")
        else:
            os.startfile('\\\\s-grp\\grp\\RADIOPHY\\Contrôle Qualité RTE\\Contrôle Qualité RTE-accélérateurs\\10_CLINAC iX 2\\10-3 CQ -EN\\10-1_CQ_quotidien\\CQ quotidien Dynalog PFROTAT - iX2.xlsm')
            os.system("pause")
    else:
        print("\n\nRESULTATS CONFORMES\n\n")
        os.system("pause")
