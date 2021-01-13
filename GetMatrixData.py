# -*- coding: utf-8 -*-

# Analyse txt file from matrix for matrix profil to send to excel file
# Author : Aurélien Corroyer-Dulmont
# Version : 01 January 2021


from tkinter import *
from tkinter.filedialog import askdirectory
from tkinter import ttk

import datetime
import codecs
import statistics
import pandas as pad 
from openpyxl import load_workbook
import win32com.client
from path import Path
import shutil
import os



'''
UI
'''
root = Tk()

root.title("Get Matrix Profils")

lbl1 = Label(root, text='Nom de la machine :')
lbl1.grid(column=0,row=0)
lbl1.config(width=40)

combo1_variable = StringVar()
combo1_values = ['Artiste', 'Clinac', 'RapidArc_iX_1', 'RapidArc_iX_2']

combo1 = ttk.Combobox(root, values= combo1_values, textvariable=combo1_variable)
combo1.grid(column=0,row=1)
combo1.config(height=5)
combo1.current(0)

lbl2 = Label(root, text='Date du CQ (au format aaaa_mm_jj) :')
lbl2.grid(column=0,row=3)
lbl2.config(width=40)

v = StringVar()
textbox1 = Entry(root, textvariable=v)
textbox1.grid(column=0, row=4)
textbox1.config(width=40)

def Quit():
	root.destroy()

butt1 = Button(root, text = 'Transfert des profils matrix', command = Quit)
butt1.grid(column=0, row=6)
butt1.config(width=20)

root.mainloop()

machineName = combo1_variable.get()
dateCQ = v.get()


def Main(machineName, dateCQ):
	date = datetime.datetime.now()
	year = date.year

	global filesOfInterest
	filesOfInterest = FileFinder(machineName, dateCQ)

	if machineName == "Artiste" or machineName == "Clinac" or machineName == "RapidArc_iX_1":
		nameProfil = ["X6_0", "X6_30", "X6_45", "X6_60", "X18_0", "X18_30", "X18_45", "X18_60"]
	elif machineName == "RapidArc_iX_2":
		nameProfil = ["X6_0", "X6_30", "X6_45", "X6_60", "X10_0", "X10_30", "X10_45", "X10_60"]

	profilToExport = pad.Series(dtype="float64")
	for elm in nameProfil:
		profilToExport[elm] = DataFinder(filesOfInterest, elm)
		profilToExport.append(DataFinder(filesOfInterest, elm))

	ExportToExcel(profilToExport, machineName)


def FileFinder(machineName, dateCQ):
	#date = datetime.datetime.now()
	#year = date.year
	annee = str(dateCQ)
	annee = annee[:4]
	fileList = []

	if machineName == "Artiste":
		for f in Path('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/0_ARTISTE/0-3_CQ-EN/0-3_CQ_mensuel/0-2 CQ_faisceaux/FICHIERS MatriXX/Profils Photons MATRIXX/' + str(annee)).walkfiles(): 
			fileList.append(f)
	if machineName == "Clinac":
		for f in Path('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/2_CLINAC/2-3 CQ -EN/2-3 CQ_mensuel/2-2 CQ_faisceaux/FICHIERS MatriXX/Profils Photons MATRIXX/' + str(annee)).walkfiles(): 
			fileList.append(f)
	if machineName == "RapidArc_iX_1":
		for f in Path('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/7_CLINAC iX 1/7-3 CQ -EN/7-3 CQ_mensuel/FICHIERS MatriXX/Profils Photons MATRIXX/' + str(annee)).walkfiles(): 
			fileList.append(f)
	if machineName == "RapidArc_iX_2":
		for f in Path('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/10_CLINAC iX 2/10-3 CQ -EN/10-3_CQ_mensuel/FICHIERS MatriXX/Profils Photons MATRIXX/' + str(annee)).walkfiles(): 
			fileList.append(f)	

	newFileList = []
	lastFileList = []
	for i in range(len(fileList)):
		newFileList.append(fileList[i].replace('Path(',''))
		lastFileList.append(newFileList[i].replace('\\','/'))

	fileList = lastFileList

	global listFilesOfInterest
	listFilesOfInterest = []
	for elm in fileList:
		if machineName == "Artiste":
			if elm[175:185] == str(dateCQ):
				listFilesOfInterest.append(elm)

		if machineName == "Clinac":
			if elm[175:185] == str(dateCQ):
				listFilesOfInterest.append(elm)

		if machineName == "RapidArc_iX_1":
			if elm[163:173] == str(dateCQ):
				listFilesOfInterest.append(elm)

		if machineName == "RapidArc_iX_2":
			if elm[166:176] == str(dateCQ):
				listFilesOfInterest.append(elm)

	### Will return the 6 matrix profils with X6 from without edge and then with wedge 30 to 60 and after X10 or X18###
	return listFilesOfInterest


def DataFinder(filesOfInterest, ProfilType):

	if ProfilType == "X6_0":
		fileProfil = pad.read_fwf(filesOfInterest[0], header = None)
	elif ProfilType == "X6_30":
		fileProfil = pad.read_fwf(filesOfInterest[1], header = None)
	elif ProfilType == "X6_45":
		fileProfil = pad.read_fwf(filesOfInterest[2], header = None)
	elif ProfilType == "X6_60":
		fileProfil = pad.read_fwf(filesOfInterest[3], header = None)
	elif ProfilType == "X10_0":
		fileProfil = pad.read_fwf(filesOfInterest[4], header = None)
	elif ProfilType == "X10_30":
		fileProfil = pad.read_fwf(filesOfInterest[5], header = None)
	elif ProfilType == "X10_45":
		fileProfil = pad.read_fwf(filesOfInterest[6], header = None)
	elif ProfilType == "X10_60":
		fileProfil = pad.read_fwf(filesOfInterest[7], header = None)
	elif ProfilType == "X18_0":
		fileProfil = pad.read_fwf(filesOfInterest[4], header = None)
	elif ProfilType == "X18_30":
		fileProfil = pad.read_fwf(filesOfInterest[5], header = None)
	elif ProfilType == "X18_45":
		fileProfil = pad.read_fwf(filesOfInterest[6], header = None)
	elif ProfilType == "X18_60":
		fileProfil = pad.read_fwf(filesOfInterest[7], header = None)


	listProfil = fileProfil[0][31:63]

	xProfil = []
	xProfilSerie = pad.Series(dtype="float64")
	for elm in listProfil:
		Temp = elm.split("\t")
		xProfil.append(float(Temp[1]))
	
	xProfilSerie[ProfilType] = xProfil

	return xProfilSerie


def ExportToExcel(profilToExport, machineName):

	if machineName == "Artiste":
		book = load_workbook('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/0_ARTISTE/0-3_CQ-EN/0-3_CQ_mensuel/0-2 CQ_faisceaux/temp_matrix_export.xlsm', read_only=False, keep_vba=True)
		writer = pad.ExcelWriter('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/0_ARTISTE/0-3_CQ-EN/0-3_CQ_mensuel/0-2 CQ_faisceaux/temp_matrix_export.xlsm', engine='openpyxl')

	if machineName == "Clinac":
		book = load_workbook('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/2_CLINAC/2-3 CQ -EN/2-3 CQ_mensuel/2-2 CQ_faisceaux/temp_matrix_export.xlsm', read_only=False, keep_vba=True)
		writer = pad.ExcelWriter('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/2_CLINAC/2-3 CQ -EN/2-3 CQ_mensuel/2-2 CQ_faisceaux/temp_matrix_export.xlsm', engine='openpyxl')
	
	if machineName == "RapidArc_iX_1":
		book = load_workbook('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/7_CLINAC iX 1/7-3 CQ -EN/7-3 CQ_mensuel/EN-0000.01-CQ_mensuel_Faisceaux_iX1.xlsm' , read_only=False, keep_vba=True)
		writer = pad.ExcelWriter('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/7_CLINAC iX 1/7-3 CQ -EN/7-3 CQ_mensuel/EN-0000.01-CQ_mensuel_Faisceaux_iX1.xlsm' , engine='openpyxl')

	if machineName == "RapidArc_iX_2":
		book = load_workbook('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/10_CLINAC iX 2/10-3 CQ -EN/10-3_CQ_mensuel/EN-0000.01-CQ_mensuel_Faisceaux_iX2.xlsm' , read_only=False, keep_vba=True)
		writer = pad.ExcelWriter('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/10_CLINAC iX 2/10-3 CQ -EN/10-3_CQ_mensuel/EN-0000.01-CQ_mensuel_Faisceaux_iX2.xlsm' , engine='openpyxl')


	writer.book = book
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
	ws = writer.sheets['ImportMatriXX']

	i=0
	for elm in profilToExport["X6_0"].values[0]:
		ws["B"+ str(8+i)] = float(profilToExport["X6_0"].values[0][i])
		i += 1

	i=0
	for elm in profilToExport["X6_30"].values[0]:
		ws["C"+ str(8+i)] = float(profilToExport["X6_30"].values[0][i])
		i += 1

	i=0
	for elm in profilToExport["X6_45"].values[0]:
		ws["D"+ str(8+i)] = float(profilToExport["X6_45"].values[0][i])
		i += 1

	i=0
	for elm in profilToExport["X6_60"].values[0]:
		ws["E"+ str(8+i)] = float(profilToExport["X6_60"].values[0][i])
		i += 1
	

	if machineName == "Artiste" or machineName == "Clinac" or machineName == "RapidArc_iX_1":
		i=0
		for elm in profilToExport["X18_0"].values[0]:
			ws["F"+ str(8+i)] = float(profilToExport["X18_0"].values[0][i])
			i += 1

		i=0
		for elm in profilToExport["X18_30"].values[0]:
			ws["G"+ str(8+i)] = float(profilToExport["X18_30"].values[0][i])
			i += 1

		i=0
		for elm in profilToExport["X18_45"].values[0]:
			ws["H"+ str(8+i)] = float(profilToExport["X18_45"].values[0][i])
			i += 1

		i=0
		for elm in profilToExport["X18_60"].values[0]:
			ws["I"+ str(8+i)] = float(profilToExport["X18_60"].values[0][i])
			i += 1


	if machineName == "RapidArc_iX_2":
		i=0
		for elm in profilToExport["X10_0"].values[0]:
			ws["F"+ str(8+i)] = float(profilToExport["X10_0"].values[0][i])
			i += 1

		i=0
		for elm in profilToExport["X10_30"].values[0]:
			ws["G"+ str(8+i)] = float(profilToExport["X10_30"].values[0][i])
			i += 1

		i=0
		for elm in profilToExport["X10_45"].values[0]:
			ws["H"+ str(8+i)] = float(profilToExport["X10_45"].values[0][i])
			i += 1

		i=0
		for elm in profilToExport["X10_60"].values[0]:
			ws["I"+ str(8+i)] = float(profilToExport["X10_60"].values[0][i])
			i += 1

	writer.save()


### Launch the script ###
Main(machineName, dateCQ)

if machineName == "Artiste":
	os.startfile('\\\\s-grp\\grp\\RADIOPHY\\Contrôle Qualité RTE\\Contrôle Qualité RTE-accélérateurs\\0_ARTISTE\\0-3_CQ-EN\\0-3_CQ_mensuel\\0-2 CQ_faisceaux\\temp_matrix_export.xlsm')
if machineName == "Clinac":
	os.startfile('\\\\s-grp\\grp\\RADIOPHY\\Contrôle Qualité RTE\\Contrôle Qualité RTE-accélérateurs\\2_CLINAC\\2-3 CQ -EN\\2-3 CQ_mensuel\\2-2 CQ_faisceaux\\temp_matrix_export.xlsm')
if machineName == "RapidArc_iX_1":
	os.startfile('\\\\s-grp\\grp\\RADIOPHY\\Contrôle Qualité RTE\\Contrôle Qualité RTE-accélérateurs\\7_CLINAC iX 1\\7-3 CQ -EN\\7-3 CQ_mensuel\\EN-0000.01-CQ_mensuel_Faisceaux_iX1.xlsm')
if machineName == "RapidArc_iX_2":
	os.startfile('\\\\s-grp\\grp\\RADIOPHY\\Contrôle Qualité RTE\\Contrôle Qualité RTE-accélérateurs\\10_CLINAC iX 2\\10-3 CQ -EN\\10-3_CQ_mensuel\\EN-0000.01-CQ_mensuel_Faisceaux_iX2.xlsm')

os.startfile('\\\\s-grp\\grp\\RADIOPHY\\Personnel\\Aurélien Corroyer-Dulmont\\Python_programmation\\pascal_prog.jpg')
