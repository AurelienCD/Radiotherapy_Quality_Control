# -*- coding: utf-8 -*-


# Analyse des dynalogs CQ mensuel du module Bras et MLC SPEED des R1-R2 : 
# => Evaluation de la différence entre la valeur attendue et réelle de 3 paramètres : 1) Position du bras, 2) Vitesse de rotation et 3) Débit de dose
# Auteur : Aurélien Corroyer-Dulmont
# Version : 13 Juillet 2020


import os, string
from math import *
import time
import datetime
import codecs
import tkinter
from tkinter.filedialog import *
from statistics import mean
from path import Path
import shutil
import pandas as pad 
from openpyxl import load_workbook
import win32com.client


def Dynalogs_BRAS_Gantry_analyser(filepath):
	"""Fonction permettant d'obtenir la position du gantry toute les 50ms (point de contrôle dynalogs) et pour les points d'index (tout les 750ms)"""

	### Analyse the number of line which is different from dynalog file from an other ###
	file = open(filepath, 'r')

	LineNotEmpty = [1]
	LineCount = 0

	while (not LineNotEmpty) != True:
		LineNotEmpty = file.readlines(1)
		LineCount += 1

	LineCount -= 1 # because the loop goes at the end + 1 line
	LineCount -= 6 # six first line are information on ARC not dynalog measure
	file.close()

	LenghtFilePath = len(filepath)
	NumeroRapid = filepath[LenghtFilePath-73]
	MachineName = "RapidArc_iX_" + str(NumeroRapid)
	AcquisitionDate = filepath[LenghtFilePath-28:LenghtFilePath-20]
	AcquisitionDateExcel = str(AcquisitionDate[6:]) + "/" + str(AcquisitionDate[4:6]) + "/" + str(AcquisitionDate[:4])

	file = open(filepath, 'r')

	i = 0
	for i in range(6):
		file.readlines(1)   
		
	# déclaration de la list qui contiendra les valeurs des positions de gantry (par point d'index et par point de contrôle dynalogs)
	GantryPositionAll = []
	

	i = 0 #vérifier que vraiment besoin de déclarer i ici
	for i in range(LineCount):
		LineRaw = file.readlines(1) #première ligne d'intérêt avec les valeurs de positions de lames, mais copie dans une list tout le premier paragraphe, donc pas utilisable
		LineListStr = ",".join(LineRaw) # passage en mode string pour ensuite repasser en mode list mais en séparant les chiffres par la virgule de façon à avoir une liste avec une donnée par ...je ne connais pas le terme
		LineListGood = LineListStr.split(",") #permet d'avoir les valeurs du premier paragraphe dans une list
		GantryPositionAll.append(int(LineListGood[0])) #permet de stocker première valeur de la ligne (position gantry) dans une list
	
	TableTempsPtIndex=[]
	TableTempsPtIndex.append(0)
	TableTempsPtIndex.append(44)
	TableTempsPtIndex.append(374)
	TableTempsPtIndex.append(414)
	TableTempsPtIndex.append(579)
	TableTempsPtIndex.append(619)
	TableTempsPtIndex.append(729)
	TableTempsPtIndex.append(769)
	TableTempsPtIndex.append(852)
	TableTempsPtIndex.append(892)
	TableTempsPtIndex.append(958)
	TableTempsPtIndex.append(998)
	TableTempsPtIndex.append(1058)
	TableTempsPtIndex.append(1098)
	TableTempsPtIndex.append(1158)
	TableTempsPtIndex.append(int(LineCount)-1)

	RealGantryPos = []
	for elm in TableTempsPtIndex:
		RealGantryPos.append(GantryPositionAll[elm])

		#INFO : RealGantryPos est maintenant une liste contenant les positions de gantry mais en 1/10ème de degrès en partant de 0 or nous on veut en degrès et en partant de 170°

		# Boucle permettant de passer la liste contenant les positions en degrès et non en 1/10ème de degrès relatif
	for i in range(len(TableTempsPtIndex)):
		RealGantryPos[i] = RealGantryPos[i]/10 #passe en degrès
		RealGantryPos[i] = 179 - RealGantryPos[i]
		if RealGantryPos[i] < 0:
			RealGantryPos[i] = 360 + RealGantryPos[i]

	RealRotationSpeed = []
	for i in range(1,len(TableTempsPtIndex)):
		diffDegree = (GantryPositionAll[TableTempsPtIndex[i]] - GantryPositionAll[TableTempsPtIndex[i]-40])/2
		RealRotationSpeed.append(diffDegree/10)

	Result = [MachineName, str(AcquisitionDateExcel)]
	j = 2
	for i in range(1,8):
		Result.append(RealGantryPos[j])
		j+=2
	j = 2
	for i in range(7):
		Result.append(RealRotationSpeed[j])
		j+=2

	return Result


def Dynalogs_BRAS_UM_analyser(filepath):   #### A ADAPTER CAR MAINTENANT IL FAUT ECART DE DEBIT ET MAX ###
	"""Fonction permettant d'obtenir le débit de dose à partir des valeurs de fraction de dose relative du dynalogs"""


	### Analyse the number of line which is different from dynalog file from an other ###
	file = open(filepath, 'r')

	LineNotEmpty = [1]
	LineCount = 0

	while (not LineNotEmpty) != True:
	   LineNotEmpty = file.readlines(1)
	   LineCount += 1

	LineCount -= 1 # because the loop goes at the end + 1 line
	LineCount -= 6 # six first line are information on ARC not dynalog measure
	file.close()


	LenghtFilePath = len(filepath)
	NumeroRapid = filepath[LenghtFilePath-71:LenghtFilePath-70]
	MachineName = "RapidArc_iX_" + str(NumeroRapid)
	AcquisitionDate = filepath[LenghtFilePath-28:LenghtFilePath-20]
	AcquisitionDateExcel = str(AcquisitionDate[6:]) + "/" + str(AcquisitionDate[4:6]) + "/" + str(AcquisitionDate[:4])


	file = open(filepath, 'r')

	# Passe les premières lignes pas utiles du dynalogs
	i = 0
	for i in range(6):
		file.readlines(1)   
		
	# Permet de stocker les valeurs de dose (relative) sur tout les points de contrôle
	DoseAll=[]
	DoseAll.append(0)
	i = 0
	for i in range(LineCount):
		LineRaw = file.readlines(1) #première ligne d'intérêt avec les valeurs de dose relative, mais copie dans une list tout le premier paragraphe, donc pas utilisable
		LineListStr = ",".join(LineRaw) # passage en mode string pour ensuite repasser en mode list mais en séparant les chiffres par la virgule de façon à avoir une liste avec une donnée par ...je ne connais pas le terme
		LineListGood = LineListStr.split(",") #permet d'avoir les valeurs du premier paragraphe dans une list
		DoseAll.append(LineListGood[0]) #DoseAll est maintenant une liste contenant les valeurs de dose en relatif (de 0 à 25000), ce dernier correspondant à 100% de la dose plannifiée
	
	TableTempsPtIndex=[]
	TableTempsPtIndex.append(0)
	TableTempsPtIndex.append(44)
	TableTempsPtIndex.append(374)
	TableTempsPtIndex.append(414)
	TableTempsPtIndex.append(579)
	TableTempsPtIndex.append(619)
	TableTempsPtIndex.append(729)
	TableTempsPtIndex.append(769)
	TableTempsPtIndex.append(852)
	TableTempsPtIndex.append(892)
	TableTempsPtIndex.append(958)
	TableTempsPtIndex.append(998)
	TableTempsPtIndex.append(1058)
	TableTempsPtIndex.append(1098)
	TableTempsPtIndex.append(1158)
	TableTempsPtIndex.append(int(LineCount)-1)

	FacteurMultiplicateurUM = 0.01 #correspond à la valeur permettant de passer du nombre relatif (0=>25000) à la valeur en UM (pour module Bras, total = 250 UM)

	# Boucle permettant de passer la liste contenant les doses en UM et non en relatif
	i = 0
	for i in range(LineCount+1):
		DoseAll[i] = float(DoseAll[i])*FacteurMultiplicateurUM #passe en UM
	
	# Boucle permettant passer des valeurs de dose à chaque point de contrôle à un débit de dose point par point
	DoseRate=[]
	DoseRate.append(0)
	j = 1
	for i in range(LineCount): 
		DoseRate.append(float(DoseAll[j])-float(DoseAll[j-1]))
		j += 1
	
	# Boucle permettant de rentrer les valeurs de dose moyenne (sur 40 pts de contôle soit 2sec) et de sélectionner celles qui correspondent à celles du tableau attendues
	DoseMean=[]
	DoseMean.append(0)
	RealDoseRate=[]
	RealDoseRate.append(0)
	j = 1
	for i in range(15): #=15 et pas 16 car la première valeur est forcement 0 car c'est une différence par rapport au point d'index précédent
		DoseMean = mean(DoseRate[TableTempsPtIndex[j]-40:TableTempsPtIndex[j]])
		DoseMean = DoseMean/50 #pour avoir valeur en dose par msec
		DoseMean = DoseMean*60000 #pour avoir valeur en dose par min
		DoseMean = round(DoseMean,3)
		RealDoseRate.append(DoseMean)
		j += 1

	Result = [MachineName, str(AcquisitionDateExcel)]
	j = 2
	for i in range(7):
		Result.append(RealDoseRate[j])
		j+=2

	return Result


def Dynalogs_MLCSPEED_UM_analyser(filepath):
	"""Fonction permettant d'obtenir le débit de dose à partir des valeurs de fraction de dose relative du dynalogs"""


	### Analyse the number of line which is different from dynalog file from an other ###
	file = open(filepath, 'r')

	LineNotEmpty = [1]
	LineCount = 0

	while (not LineNotEmpty) != True:
	   LineNotEmpty = file.readlines(1)
	   LineCount += 1

	LineCount -= 1 # because the loop goes at the end + 1 line
	LineCount -= 6 # six first line are information on ARC not dynalog measure
	file.close()


	LenghtFilePath = len(filepath)
	NumeroRapid = filepath[LenghtFilePath-60]
	MachineName = "RapidArc_iX_" + str(NumeroRapid)
	AcquisitionDate = filepath[LenghtFilePath-28:LenghtFilePath-20]
	AcquisitionDateExcel = str(AcquisitionDate[6:]) + "/" + str(AcquisitionDate[4:6]) + "/" + str(AcquisitionDate[:4])


	file = open(filepath, 'r')

	# Passe les premières lignes pas utiles du dynalogs
	i = 0
	for i in range(6):
		file.readlines(1)   
		
	# Permet de stocker les valeurs de dose (relative) sur tout les points de contrôle
	DoseAll=[]
	DoseAll.append(0)
	i = 0
	for i in range(LineCount):
		LineRaw = file.readlines(1) #première ligne d'intérêt avec les valeurs de dose relative, mais copie dans une list tout le premier paragraphe, donc pas utilisable
		LineListStr = ",".join(LineRaw) # passage en mode string pour ensuite repasser en mode list mais en séparant les chiffres par la virgule de façon à avoir une liste avec une donnée par ...je ne connais pas le terme
		LineListGood = LineListStr.split(",") #permet d'avoir les valeurs du premier paragraphe dans une list
		DoseAll.append(LineListGood[0]) #DoseAll est maintenant une liste contenant les valeurs de dose en relatif (de 0 à 25000), ce dernier correspondant à 100% de la dose plannifiée
	

	TableTempsPtIndex=[]
	TableTempsPtIndex.append(97)
	TableTempsPtIndex.append(194)
	TableTempsPtIndex.append(338)
	TableTempsPtIndex.append(int(LineCount)-1)

	FacteurMultiplicateurUM = 0.0072 #correspond à la valeur permettant de passer du nombre relatif (0=>25000) à la valeur en UM (pour module MLC SPEED, total = 180 UM)

	# Boucle permettant de passer la liste contenant les doses en UM et non en relatif
	i = 0
	for i in range(LineCount+1):
		DoseAll[i] = float(DoseAll[i])*FacteurMultiplicateurUM #passe en UM
	
	# Boucle permettant passer des valeurs de dose à chaque point de contrôle à un débit de dose point par point
	DoseRate=[]
	DoseRate.append(0)
	for i in range(1,LineCount+1):
		DoseRate.append(float(DoseAll[i])-float(DoseAll[i-1]))
	
	# Boucle permettant de rentrer les valeurs de dose moyenne (sur 40 pts de contôle soit 2sec) et de sélectionner celles qui correspondent à celles du tableau attendues
	RealDoseRate=[]
	j = 0
	for i in range(4): 
		DoseMean = mean(DoseRate[TableTempsPtIndex[j]-40:TableTempsPtIndex[j]])
		DoseMean = DoseMean/50 #pour avoir valeur en dose par msec
		DoseMean = DoseMean*60000 #pour avoir valeur en dose par min
		DoseMean = round(DoseMean,3)
		RealDoseRate.append(DoseMean)
		j += 1

	Result = [MachineName, str(AcquisitionDateExcel)]
	for elm in RealDoseRate:
		Result.append(elm)

	return Result



######### Create Pandas Excel functions to upload the results in the excel data base file ###########
def ExportToExcel(Module, ListOfResults):
	MachineName = str(ListOfResults[0])
	ListOfResultsToExcel = ListOfResults[1:]
	if MachineName == "RapidArc_iX_1":
		book = load_workbook('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/7_CLINAC iX 1/7-3 CQ -EN/7-3 CQ_mensuel/EN-0000.01-CQ_mensuel_MLCdyn Dynalog_iX1.xlsm', read_only=False, keep_vba=True)
		writer = pad.ExcelWriter('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/7_CLINAC iX 1/7-3 CQ -EN/7-3 CQ_mensuel/EN-0000.01-CQ_mensuel_MLCdyn Dynalog_iX1.xlsm', engine='openpyxl') 
		writer.book = book
		writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
		ws = writer.sheets[str(Module)]
		if Module == "BRAS_gantry":
			ws["D12"] = str(ListOfResultsToExcel[0])
			ws["E12"] = float(ListOfResultsToExcel[1])
			ws["I12"] = float(ListOfResultsToExcel[2])
			ws["M12"] = float(ListOfResultsToExcel[3])
			ws["Q12"] = float(ListOfResultsToExcel[4])
			ws["U12"] = float(ListOfResultsToExcel[5])
			ws["Y12"] = float(ListOfResultsToExcel[6])
			ws["AC12"] = float(ListOfResultsToExcel[7])
			ws["G12"] = float(ListOfResultsToExcel[8])
			ws["K12"] = float(ListOfResultsToExcel[9])
			ws["O12"] = float(ListOfResultsToExcel[10])
			ws["S12"] = float(ListOfResultsToExcel[11])
			ws["W12"] = float(ListOfResultsToExcel[12])
			ws["AA12"] = float(ListOfResultsToExcel[13])
			ws["AE12"] = float(ListOfResultsToExcel[14])
		
		elif Module == "BRAS_UM":
			ws["D12"] = str(ListOfResultsToExcel[0])
			ws["E12"] = float(ListOfResultsToExcel[1])
			ws["G12"] = float(ListOfResultsToExcel[2])
			ws["I12"] = float(ListOfResultsToExcel[3])
			ws["K12"] = float(ListOfResultsToExcel[4])
			ws["M12"] = float(ListOfResultsToExcel[5])
			ws["O12"] = float(ListOfResultsToExcel[6])
			ws["Q12"] = float(ListOfResultsToExcel[7])

		else:
			ws["D12"] = str(ListOfResultsToExcel[0])
			ws["E12"] = float(ListOfResultsToExcel[1])
			ws["G12"] = float(ListOfResultsToExcel[2])
			ws["I12"] = float(ListOfResultsToExcel[3])
			ws["K12"] = float(ListOfResultsToExcel[4])

		writer.save()
		xl = win32com.client.Dispatch('Excel.Application')
		xl.Workbooks.Open(Filename = '//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/7_CLINAC iX 1/7-3 CQ -EN/7-3 CQ_mensuel/EN-0000.01-CQ_mensuel_MLCdyn Dynalog_iX1.xlsm', ReadOnly=1)  
		xl.Worksheets(str(Module)).Activate()
		xl.Application.Run("Archivage_" + str(Module))
		xl.Application.Quit()
		del xl

	else:
		book = load_workbook('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/10_CLINAC iX 2/10-3 CQ -EN/10-3_CQ_mensuel/EN-0000.01-CQ_mensuel_MLCdyn Dynalog_iX2.xlsm', read_only=False, keep_vba=True)
		writer = pad.ExcelWriter('//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/10_CLINAC iX 2/10-3 CQ -EN/10-3_CQ_mensuel/EN-0000.01-CQ_mensuel_MLCdyn Dynalog_iX2.xlsm', engine='openpyxl') 
		writer.book = book
		writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
		ws = writer.sheets[str(Module)]
		if Module == "BRAS_gantry":
			ws["D12"] = str(ListOfResultsToExcel[0])
			ws["E12"] = float(ListOfResultsToExcel[1])
			ws["I12"] = float(ListOfResultsToExcel[2])
			ws["M12"] = float(ListOfResultsToExcel[3])
			ws["Q12"] = float(ListOfResultsToExcel[4])
			ws["U12"] = float(ListOfResultsToExcel[5])
			ws["Y12"] = float(ListOfResultsToExcel[6])
			ws["AC12"] = float(ListOfResultsToExcel[7])
			ws["G12"] = float(ListOfResultsToExcel[8])
			ws["K12"] = float(ListOfResultsToExcel[9])
			ws["O12"] = float(ListOfResultsToExcel[10])
			ws["S12"] = float(ListOfResultsToExcel[11])
			ws["W12"] = float(ListOfResultsToExcel[12])
			ws["AA12"] = float(ListOfResultsToExcel[13])
			ws["AE12"] = float(ListOfResultsToExcel[14])
		
		elif Module == "BRAS_UM":
			ws["D12"] = str(ListOfResultsToExcel[0])
			ws["E12"] = float(ListOfResultsToExcel[1])
			ws["G12"] = float(ListOfResultsToExcel[2])
			ws["I12"] = float(ListOfResultsToExcel[3])
			ws["K12"] = float(ListOfResultsToExcel[4])
			ws["M12"] = float(ListOfResultsToExcel[5])
			ws["O12"] = float(ListOfResultsToExcel[6])
			ws["Q12"] = float(ListOfResultsToExcel[7])

		else:
			ws["D12"] = str(ListOfResultsToExcel[0])
			ws["E12"] = float(ListOfResultsToExcel[1])
			ws["G12"] = float(ListOfResultsToExcel[2])
			ws["I12"] = float(ListOfResultsToExcel[3])
			ws["K12"] = float(ListOfResultsToExcel[4])

		writer.save()
		xl = win32com.client.Dispatch('Excel.Application')
		xl.Workbooks.Open(Filename = '//s-grp/grp/RADIOPHY/Contrôle Qualité RTE/Contrôle Qualité RTE-accélérateurs/10_CLINAC iX 2/10-3 CQ -EN/10-3_CQ_mensuel/EN-0000.01-CQ_mensuel_MLCdyn Dynalog_iX2.xlsm', ReadOnly=1)  
		xl.Worksheets(str(Module)).Activate()
		xl.Application.Run("Archivage_" + str(Module))
		xl.Application.Quit()
		del xl


### To get the path of the dynalog files of each module ###
fileList = []
for f in Path('Z:/Aurelien_Dynalogs/0_CQ Mensuel MLC Dyn').walkfiles(): 
	fileList.append(f)

temp_BRAS_Gantry_File_List = []
temp_BRAS_UM_File_List = []
temp_MLCSPEED_UM_File_List = []

for elm in fileList:
	if elm[70] == "2":
		temp_BRAS_Gantry_File_List.append(elm)
	elif elm[70] == "3":
		temp_BRAS_UM_File_List.append(elm)
	elif elm[70] == "4":
		temp_MLCSPEED_UM_File_List.append(elm)

### We select half of the list as we only want bench A for UM purpose ###
BRAS_Gantry_File_List = []
for elm in temp_BRAS_Gantry_File_List:
	if elm[95] =="A":
		BRAS_Gantry_File_List.append(elm)

BRAS_UM_File_List = []
for elm in temp_BRAS_UM_File_List:
	if elm[93] =="A":
		BRAS_UM_File_List.append(elm)

MLCSPEED_UM_File_List = []
for elm in temp_MLCSPEED_UM_File_List:
	if elm[82] =="A":
		MLCSPEED_UM_File_List.append(elm)


newFileList = []
lastFileList = []
for i in range(len(BRAS_Gantry_File_List)):
	newFileList.append(BRAS_Gantry_File_List[i].replace('Path(',''))
	lastFileList.append(newFileList[i].replace('\\','/'))
BRAS_Gantry_File_List = lastFileList

newFileList = []
lastFileList = []
for i in range(len(BRAS_UM_File_List)):
	newFileList.append(BRAS_UM_File_List[i].replace('Path(',''))
	lastFileList.append(newFileList[i].replace('\\','/'))
BRAS_UM_File_List = lastFileList

newFileList = []
lastFileList = []
for i in range(len(MLCSPEED_UM_File_List)):
	newFileList.append(MLCSPEED_UM_File_List[i].replace('Path(',''))
	lastFileList.append(newFileList[i].replace('\\','/'))
MLCSPEED_UM_File_List = lastFileList


print("\n\nIl y a " + str(len(BRAS_Gantry_File_List)) + " fichiers dynalogs du module "'Bras Gantry'" à analyser")
print("\n\nIl y a " + str(len(BRAS_UM_File_List)) + " fichiers dynalogs du module "'Bras UM'" à analyser")
print("\n\nIl y a " + str(len(MLCSPEED_UM_File_List)) + " fichiers dynalogs du module "'MLC SPEED UM'" à analyser")


if len(BRAS_Gantry_File_List) != 0 or len(BRAS_UM_File_List) != 0 or len(MLCSPEED_UM_File_List) != 0:
	print("\n\nLancement du programme d'analyse ...\n\n")
	for i in range(len(BRAS_Gantry_File_List)):
		ListOfResultsBRAS_Gantry = Dynalogs_BRAS_Gantry_analyser(str(BRAS_Gantry_File_List[i]))
		ExportToExcel("BRAS_gantry", ListOfResultsBRAS_Gantry)

	for i in range(len(BRAS_UM_File_List)):
		ListOfResultsBRAS_UM = Dynalogs_BRAS_UM_analyser(str(BRAS_UM_File_List[i]))
		ExportToExcel("BRAS_UM", ListOfResultsBRAS_UM)

	for i in range(len(MLCSPEED_UM_File_List)):
		ListOfResultsMLCSPEED_UM = Dynalogs_MLCSPEED_UM_analyser(str(MLCSPEED_UM_File_List[i]))
		ExportToExcel("MLC_Speed_UM", ListOfResultsMLCSPEED_UM)
	
	print("\n\nANALYSE TERMINEE\n\n")

	#### Déplacement des fichiers dynalogs analysés dans un répertoire d'archive ###
	for i in range(len(fileList)):
		newFileList.append(fileList[i].replace('Path(',''))
		lastFileList.append(newFileList[i].replace('\\','/'))
	ExportFileList = lastFileList

	for file in ExportFileList:
		shutil.move(file[:69], 'Z:/Aurelien_Dynalogs/Results_Analyses_Dynalogs/CQ Mensuel MLC Dyn/Archive')


	os.startfile('\\\\s-grp\\grp\\RADIOPHY\\Contrôle Qualité RTE\\Contrôle Qualité RTE-accélérateurs\\7_CLINAC iX 1\\7-3 CQ -EN\\7-3 CQ_mensuel\\EN-0000.01-CQ_mensuel_MLCdyn Dynalog_iX1.xlsm')
	os.startfile('\\\\s-grp\\grp\\RADIOPHY\\Contrôle Qualité RTE\\Contrôle Qualité RTE-accélérateurs\\10_CLINAC iX 2\\10-3 CQ -EN\\10-3_CQ_mensuel\\EN-0000.01-CQ_mensuel_MLCdyn Dynalog_iX2.xlsm')
	os.system("pause")


